from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ...schemas.grading_response import GradingResponse


def build_workbook(results: list[GradingResponse]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ket qua"
    headers = [
        "STT",
        "Họ và tên",
        "Lớp",
        "Môn thi",
        "Mã đề",
        *[f"Câu {index}" for index in range(1, 41)],
        "Số câu đúng",
        "Số câu sai",
        "Số câu bỏ trống",
        "Số câu tô nhiều",
        "Số câu cần kiểm tra",
        "Điểm",
        "Trạng thái",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2F0D9")

    for index, result in enumerate(results, start=1):
        answers = {answer.question: answer.detected_answer for answer in result.answers}
        sheet.append(
            [
                index,
                result.student.name,
                result.student.class_name,
                result.student.subject,
                result.student.exam_code,
                *[answers.get(question, "") for question in range(1, 41)],
                result.correct_count,
                result.wrong_count,
                result.blank_count,
                result.multiple_count,
                result.review_count,
                result.score,
                result.status,
            ]
        )

    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(28, max(10, max(len(str(cell.value or "")) for cell in column) + 2))

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream
