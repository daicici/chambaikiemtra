from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def export_results_to_excel(results: list[dict], output_path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Ket qua cham bai"
    headers = [
        "STT",
        "Họ và tên",
        "Lớp",
        "Môn thi",
        "Mã đề",
        *[f"Câu {index}" for index in range(1, 41)],
        "Số câu đúng",
        "Số câu sai",
        "Số câu bỏ trống",
        "Số câu tô nhiều",
        "Số câu cần kiểm tra",
        "Điểm",
        "Trạng thái phiếu",
    ]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="D9EAD3")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, result in enumerate(results, start=1):
        answers = {item["question"]: item["selected"] for item in result["answers"]}
        sheet.append(
            [
                row_index,
                result["student_name"]["normalized"],
                result["class_name"]["normalized"],
                result["subject"]["normalized"],
                result["exam_code"]["normalized"],
                *[answers.get(index, "") for index in range(1, 41)],
                result["correct_count"],
                result["wrong_count"],
                result["blank_count"],
                result["multiple_count"],
                result["review_count"],
                result["score"],
                result["status"],
            ]
        )

    for column in sheet.columns:
        width = min(28, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[column[0].column_letter].width = width

    workbook.save(output_path)
    return output_path
